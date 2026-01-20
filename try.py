from openpyxl.styles import Protection
from openpyxl.worksheet.table import Table, TableStyleInfo

def protect_sheet_column(ws, unlock_col_letters="Q"):
    max_row = ws.max_row
    max_col = ws.max_column

    # 1️⃣ Lock everything
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.protection = Protection(locked=True)

    # 2️⃣ Unlock editable column (Q)
    unlock_col_idx = ord(unlock_col_letters.upper()) - ord("A") + 1
    for row in ws.iter_rows(min_row=1, max_row=max_row,
                            min_col=unlock_col_idx, max_col=unlock_col_idx):
        for cell in row:
            cell.protection = Protection(locked=False)

    # 3️⃣ Unlock header rows where Column A == "Added"
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        if row[0].value == "Added":
            for cell in row:
                cell.protection = Protection(locked=False)

    # 4️⃣ Convert used range to Table (filters auto-enabled)
    end_col_letter = ws.cell(row=1, column=max_col).column_letter
    table_ref = f"A1:{end_col_letter}{max_row}"

    table = Table(displayName="AddedTable", ref=table_ref)

    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    ws.add_table(table)

    # 5️⃣ Protect sheet (filters still usable)
    ws.protection.enableAutoFilter = True
    ws.protection.enable()
    ws.protection.sheet = True
